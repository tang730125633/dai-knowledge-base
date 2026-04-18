#!/usr/bin/env python3
"""
用戴总 2026-04-18 新整理的"副本文件库.xlsx"（1646 条）重新做权威去重

任务（Tang 明确 2026-04-18 12:30）：
1. 比对戴总新清单 vs 我方 samr 2020 条
2. 重复的 → 生成"需要从 samr 删除"清单
3. 不重复的 → 生成"推荐给戴总"Excel
4. 在我方 samr_review_index.json 里逐条标记 'in_dai_backend: true'
"""
import json
import openpyxl
import time
from pathlib import Path
from normalize_code import normalize_code
from power_filter import is_power_relevant, classify_category, classify_phase, classify_std_type

KB = Path(__file__).parent
DAI_EXCEL = Path.home() / 'Desktop' / '副本文件库.xlsx'
OUT = Path.home() / 'dai-delivery' / 'weekly-batches'


def load_dai_library_v2():
    """加载戴总新整理的文件库"""
    wb = openpyxl.load_workbook(DAI_EXCEL, data_only=True)
    ws = wb['Sheet1']
    library = {}  # normalized_code → info
    total = 0
    with_code = 0
    dup = 0

    # 第一行是表头: ['我们的文件名', 'HG T 20615-2009', '1646']
    # 第一列是文件名，第二列是标准号。跳过第 1 行
    for i in range(2, ws.max_row + 1):
        total += 1
        filename = ws.cell(i, 1).value or ''
        raw_code = ws.cell(i, 2).value or ''
        raw_code = str(raw_code).strip()
        if not raw_code:
            continue

        with_code += 1
        n = normalize_code(raw_code)
        if not n:
            continue

        if n in library:
            dup += 1
            # 同一编号多个文件，记录所有
            existing = library[n].setdefault('all_filenames', [library[n]['filename']])
            if filename and filename not in existing:
                existing.append(filename)
            existing_raw = library[n].setdefault('all_raw_codes', [library[n]['raw_code']])
            if raw_code not in existing_raw:
                existing_raw.append(raw_code)
        else:
            library[n] = {
                'raw_code': raw_code,
                'filename': filename,
                'original_row': i,
            }
    print(f"✓ 戴总新清单 加载完成")
    print(f"  总行数: {total}")
    print(f"  有编号: {with_code}")
    print(f"  规范化唯一编号: {len(library)}")
    print(f"  同编号多文件: {dup}")
    return library


def load_samr():
    data = json.loads((KB / '_crawl_data' / 'samr_review_index.json').read_text())
    return data.get('items', [])


def main():
    OUT.mkdir(exist_ok=True)

    # 1. 加载戴总新清单
    dai = load_dai_library_v2()
    dai_norm = set(dai.keys())

    # 2. 加载我方 samr
    samr = load_samr()
    samr_by_norm = {}
    for item in samr:
        n = normalize_code(item.get('code', ''))
        if n:
            samr_by_norm[n] = item
    samr_norm = set(samr_by_norm.keys())

    print(f"\n✓ 我方 samr 数据")
    print(f"  原始条目: {len(samr)}")
    print(f"  规范化唯一: {len(samr_norm)}")

    # 3. 查重分析
    overlap = dai_norm & samr_norm
    only_dai = dai_norm - samr_norm
    only_samr = samr_norm - dai_norm

    print(f"\n{'='*60}")
    print(f"📊 查重结果")
    print(f"{'='*60}")
    print(f"戴总新清单 1646 行 → 规范化 {len(dai_norm)} 个唯一编号")
    print(f"我方 samr 2020 条 → 规范化 {len(samr_norm)} 个唯一编号")
    print()
    print(f"🔄 重复（我方要删的）: {len(overlap)} 条")
    print(f"⬇️  仅戴总有（我们没爬到）: {len(only_dai)} 条")
    print(f"🟢 仅我方有（可推荐给戴总）: {len(only_samr)} 条")

    # 4. 生成"推荐给戴总"Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '推荐给戴总的新标准'

    headers = ['标准编号', '标题', '类型', '大类', '子分类', '工程环节', '详情页 URL', '备注']
    ws.append(headers)

    # 设置列宽
    widths = [22, 55, 10, 12, 18, 10, 60, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # 表头样式
    from openpyxl.styles import Font, PatternFill
    bold = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='4472C4')
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill

    import re
    # 正式标准号模式（开头必须是这些前缀 + 数字）
    FORMAL_PREFIX = re.compile(r'^(GB(/T)?|DL(/T)?|NB(/T)?|SL(/T)?|JGJ(/T)?|CJJ(/T)?|JB(/T)?|YD(/T)?|Q/GDW|Q/CSG|GBT|DLT|NBT|SLT|JGJT|CJJT|JBT|YDT|QGDW|QCSG|DB\d+(/T)?|HG(/T)?|HGT)[\s-]*\d', re.IGNORECASE)

    rows_added = 0
    rejected_by_filter = 0
    rejected_by_code_format = 0
    rejected_by_status = 0

    # 先拆出：正式已发布的 vs 立项号
    for norm_code in sorted(only_samr):
        item = samr_by_norm[norm_code]
        title = item.get('title', '')
        keyword = item.get('keyword', '')
        category = item.get('category', '')
        raw_code = item.get('code', '')
        status = item.get('status', '')

        # 过滤 1: 必须是正式标准号（GB/DL/NB/SL/JGJ/Q/GDW 等开头）
        if not FORMAL_PREFIX.match(raw_code):
            rejected_by_code_format += 1
            continue

        # 过滤 2: 淘汰已终止/作废的
        if status in ('已终止', '作废', '废止'):
            rejected_by_status += 1
            continue

        # 过滤 3: 电力/新能源关键词
        search_text = f"{title} {keyword} {category}"
        passed, score, hits = is_power_relevant(search_text)
        if not passed:
            rejected_by_filter += 1
            continue

        std_type = classify_std_type(raw_code)
        inferred_category = classify_category(search_text) or ''
        inferred_phase = classify_phase(search_text) or item.get('phase', '')

        ws.append([
            raw_code,
            title,
            std_type or '未定',
            inferred_category,
            keyword,
            inferred_phase,
            item.get('detail_url', ''),
            f"状态:{status}" if status else ''
        ])
        rows_added += 1

    print(f"\n📊 Excel 过滤详情：")
    print(f"  仅我方有总量: {len(only_samr)}")
    print(f"  - 立项号/非正式编号: {rejected_by_code_format}")
    print(f"  - 已终止/作废: {rejected_by_status}")
    print(f"  - 非电力相关: {rejected_by_filter}")
    print(f"  = 实际推荐给戴总: {rows_added}")

    out_file = OUT / '推荐给戴总的新标准.xlsx'
    wb.save(out_file)
    print(f"\n✅ 推荐 Excel 已生成：{out_file}")
    print(f"   含 {rows_added} 条（过滤掉 {rejected_by_filter} 条非电力相关）")

    # 5. 在 samr_review_index 里标记重复的
    samr_data = json.loads((KB / '_crawl_data' / 'samr_review_index.json').read_text())
    marked = 0
    for item in samr_data['items']:
        n = normalize_code(item.get('code', ''))
        if n and n in overlap:
            item['in_dai_backend'] = True
            item['dai_filename'] = dai[n].get('filename', '')
            marked += 1

    samr_data['dai_comparison'] = {
        'compared_at': time.strftime('%Y-%m-%dT%H:%M:%S'),
        'dai_source': '副本文件库.xlsx 2026-04-18',
        'dai_unique_codes': len(dai_norm),
        'overlap': len(overlap),
        'only_dai': len(only_dai),
        'only_samr': len(only_samr),
    }
    (KB / '_crawl_data' / 'samr_review_index.json').write_text(
        json.dumps(samr_data, ensure_ascii=False, indent=2))
    print(f"\n✅ samr_review_index.json 已标记 {marked} 条 in_dai_backend=true")

    # 6. 生成"戴总独有"清单（可用来做第二版爬虫目标）
    only_dai_list = sorted(only_dai)
    (OUT / 'dai-only-codes.json').write_text(json.dumps({
        'generated_at': time.strftime('%Y-%m-%dT%H:%M:%S'),
        'source': '副本文件库.xlsx',
        'total': len(only_dai_list),
        'codes': only_dai_list[:200],  # 只展示前 200
        'codes_all': only_dai_list,
    }, ensure_ascii=False, indent=2))
    print(f"\n✅ 戴总独有编号清单：{OUT / 'dai-only-codes.json'}")
    print(f"   共 {len(only_dai_list)} 条")

    # 7. 更新主数据库的戴总来源（主库放在 knowledge-base 仓库里）
    master_file = KB / '_crawl_data' / 'master-database.json'
    master = json.loads(master_file.read_text())
    for n in dai_norm:
        if n in master['standards']:
            master['standards'][n]['in_excel'] = True  # 正确标记
            master['standards'][n]['sources']['excel'] = {
                'filename': dai[n].get('filename', ''),
                'source_version': '副本文件库.xlsx 2026-04-18',
            }
    master_file.write_text(json.dumps(master, ensure_ascii=False, indent=2))
    print(f"\n✅ master-database.json 已更新戴总最新清单")


if __name__ == '__main__':
    main()
