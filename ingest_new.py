#!/usr/bin/env python3
"""
把"推荐给戴总的新标准.xlsx"里的 1207 条按 4 级目录写进仓库 .md

流程：
1. 读 Excel 的 1207 行
2. 对每行用 subcategory_mapping.resolve_target_dir() 找目标目录
3. 目录不存在 → 新建 + 生成空的 标准索引.md 骨架
4. 把这条 append 进对应 .md 表格
5. 更新文件头部"标准数量"字段

安全：
- 默认 --dry-run，只打印会改什么
- 生效：python3 ingest_new.py --apply
"""
import json
import re
import sys
import time
from pathlib import Path
from collections import defaultdict, Counter
import openpyxl

from subcategory_mapping import resolve_target_dir, EXISTING_SUBCATEGORIES

KB = Path(__file__).parent
EXCEL = Path.home() / 'dai-delivery' / 'weekly-batches' / '2026-04-W16' / '推荐给戴总的新标准.xlsx'
BATCH_DIR = Path.home() / 'dai-delivery' / 'weekly-batches' / '2026-04-W16'

SKELETON_TEMPLATE = """# {subcategory} - {phase} - {std_type}

> 品类: {category} / {subcategory}
> 生命周期: {phase}
> 标准来源: {std_type}
> 状态: ✅ 已索引 {count} 条

| 序号 | 标准编号 | 标准名称 | 资源链接 | 状态 |
|------|----------|----------|----------|------|
{rows}
> 📅 索引更新时间: {date}
"""


def load_excel_rows():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({
            'code': (ws.cell(r, 1).value or '').strip(),
            'title': (ws.cell(r, 2).value or '').strip(),
            'std_type': (ws.cell(r, 3).value or '').strip(),
            'category': (ws.cell(r, 4).value or '').strip(),
            'keyword': (ws.cell(r, 5).value or '').strip(),
            'phase': (ws.cell(r, 6).value or '').strip(),
            'detail_url': (ws.cell(r, 7).value or '').strip(),
            'note': (ws.cell(r, 8).value or '').strip(),
        })
    return rows


def resolve_rows(rows):
    """返回 {(cat_dir, sub_dir, phase_dir, type_dir): [rows]}"""
    groups = defaultdict(list)
    unresolved = []
    for row in rows:
        r = resolve_target_dir(
            keyword=row['keyword'],
            category=row['category'],
            phase=row['phase'],
            std_type=row['std_type'],
        )
        if r is None:
            unresolved.append(row)
            continue
        groups[r].append(row)
    return groups, unresolved


def read_existing_codes(md_path: Path) -> set:
    """读取 md 文件里已有的规范化编号（避免重复 ingest）"""
    if not md_path.exists():
        return set()
    from normalize_code import normalize_code
    content = md_path.read_text()
    codes = set()
    for line in content.split('\n'):
        m = re.match(r'^\|\s*\d+\s*\|\s*([^|]+?)\s*\|', line)
        if m:
            n = normalize_code(m.group(1).strip())
            if n:
                codes.add(n)
    return codes


def ingest_group(target: tuple, items: list, dry_run: bool):
    """处理一组：写入 .md"""
    cat_dir, sub_dir, phase_dir, type_dir = target
    folder = KB / cat_dir / sub_dir / phase_dir / type_dir
    md = folder / '标准索引.md'

    # 1. 文件不存在 → 新建目录 + skeleton
    created = False
    if not folder.exists():
        if not dry_run:
            folder.mkdir(parents=True, exist_ok=True)
        created = True

    # 2. 读现有 md（如果有）
    from normalize_code import normalize_code
    existing_codes = read_existing_codes(md) if md.exists() else set()

    # 3. 判断要 append 哪些（跳过已存在的）
    to_add = []
    skipped = []
    for row in items:
        n = normalize_code(row['code'])
        if n in existing_codes:
            skipped.append(row)
            continue
        to_add.append(row)
        existing_codes.add(n)  # 防本批次重复

    if not to_add:
        return {
            'path': str(md.relative_to(KB)),
            'created': created,
            'added': 0,
            'skipped_duplicates': len(skipped),
        }

    # 4. 组装 .md 新内容
    if md.exists():
        content = md.read_text()
        # 在现有表格末尾追加
        lines = content.split('\n')
        # 找最后一行以 "|" 开头的行位置
        last_table_line = -1
        for i, line in enumerate(lines):
            if line.strip().startswith('|') and not line.strip().startswith('|---'):
                last_table_line = i
        # 找现有最大序号
        max_seq = 0
        for line in lines:
            m = re.match(r'^\|\s*(\d+)\s*\|', line)
            if m:
                max_seq = max(max_seq, int(m.group(1)))
        # 生成新行
        new_rows = []
        for i, row in enumerate(to_add, 1):
            seq = max_seq + i
            status = f"📋 待核实"
            detail_url = row.get('detail_url', '')
            link_host = 'samr.gov.cn' if 'samr.gov.cn' in detail_url else 'bzfxw.com' if 'bzfxw' in detail_url else '链接'
            new_row = f"| {seq} | {row['code']} | {row['title']} | [{link_host}]({detail_url}) | {status} |"
            new_rows.append(new_row)

        # 插入到最后表格行之后
        if last_table_line >= 0:
            lines[last_table_line+1:last_table_line+1] = new_rows
            # 更新"已索引 N 条"
            for i, line in enumerate(lines):
                if '已索引' in line and '条' in line:
                    lines[i] = re.sub(r'已索引\s*\d+\s*条', f'已索引 {max_seq + len(to_add)} 条', line)
                if re.match(r'>\s*标准数量[：:]', line):
                    lines[i] = re.sub(r'标准数量[：:]\s*\d+', f'标准数量: {max_seq + len(to_add)}', line)
            new_content = '\n'.join(lines)
        else:
            # 非常异常情况：没找到表格行，全量重写（不该发生）
            new_content = content

    else:
        # 新建 md skeleton
        new_rows = []
        for i, row in enumerate(to_add, 1):
            detail_url = row.get('detail_url', '')
            link_host = 'samr.gov.cn' if 'samr.gov.cn' in detail_url else 'bzfxw.com' if 'bzfxw' in detail_url else '链接'
            new_rows.append(f"| {i} | {row['code']} | {row['title']} | [{link_host}]({detail_url}) | 📋 待核实 |")
        new_content = SKELETON_TEMPLATE.format(
            subcategory=sub_dir,
            phase=phase_dir,
            std_type=type_dir,
            category=cat_dir,
            count=len(to_add),
            rows='\n'.join(new_rows),
            date=time.strftime('%Y-%m-%d'),
        )

    if not dry_run:
        md.write_text(new_content)

    return {
        'path': str(md.relative_to(KB)),
        'created': created,
        'added': len(to_add),
        'skipped_duplicates': len(skipped),
    }


def main():
    dry_run = '--apply' not in sys.argv

    print(f"{'='*60}")
    print(f"  ingest_new.py {'(dry-run)' if dry_run else '(APPLYING)'}")
    print(f"{'='*60}\n")

    rows = load_excel_rows()
    print(f"📊 Excel 加载: {len(rows)} 行\n")

    groups, unresolved = resolve_rows(rows)
    print(f"✓ 分组完成：")
    print(f"  能映射到目标目录: {sum(len(v) for v in groups.values())} 条（{len(groups)} 个目录）")
    print(f"  未能映射（关键词未知）: {len(unresolved)} 条")

    # 执行
    reports = []
    total_added = 0
    total_skipped = 0
    dirs_created = 0

    for target, items in sorted(groups.items(), key=lambda x: (-len(x[1]), x[0])):
        r = ingest_group(target, items, dry_run=dry_run)
        reports.append(r)
        total_added += r['added']
        total_skipped += r['skipped_duplicates']
        if r['created']:
            dirs_created += 1

    # 总结
    print(f"\n{'='*60}")
    print(f"📊 执行结果")
    print(f"{'='*60}")
    print(f"  新建目录: {dirs_created}")
    print(f"  新增条目: {total_added}")
    print(f"  跳过重复: {total_skipped}")
    print(f"  未能映射（需手工）: {len(unresolved)}")

    # Top 10
    print(f"\n🔝 新增条目最多的 10 个文件:")
    for r in sorted(reports, key=lambda x: -x['added'])[:10]:
        mark = '🆕' if r['created'] else '➕'
        print(f"  {mark} {r['added']:3} 条 → {r['path']}")

    # 保存报告
    out = BATCH_DIR / ('ingest-preview.json' if dry_run else 'ingest-result.json')
    out.write_text(json.dumps({
        'mode': 'dry-run' if dry_run else 'applied',
        'generated_at': time.strftime('%Y-%m-%dT%H:%M:%S'),
        'summary': {
            'rows_in_excel': len(rows),
            'groups_resolved': len(groups),
            'groups_unresolved': len(unresolved),
            'dirs_created': dirs_created,
            'entries_added': total_added,
            'entries_skipped': total_skipped,
        },
        'unresolved_samples': unresolved[:20],
        'reports': reports,
    }, ensure_ascii=False, indent=2))
    print(f"\n💾 详细报告: {out}")

    if dry_run:
        print(f"\n⚠️  这是 dry-run。确认无误后跑：python3 ingest_new.py --apply")
    else:
        print(f"\n✅ 已应用！下一步：")
        print(f"  1. 跑 sync_pdf_folder.py 把新目录同步到 delivery/pdfs/")
        print(f"  2. git add . && git commit -m 'feat: ingest {total_added} new standards'")
        print(f"  3. git push")


if __name__ == '__main__':
    main()
