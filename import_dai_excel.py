#!/usr/bin/env python3
"""
把戴总的"文件库.xlsx"导入为规范化的 JSON

戴总 Excel 格式：
  Sheet1: 2 列
    第 1 列: 我们的文件名（PDF 文件名）
    第 2 列: 提取标准号（可能为空）

戴总明确说："没有标准号的（书籍/地方/企业文件），先不管"
所以这里只导入第 2 列非空的记录，作为查重对照数据。

输出：_crawl_data/dai_file_library.json
  {
    "规范化编号": {
      "raw_code": "原始编号（戴总的写法）",
      "filename": "对应的 PDF 文件名",
      "original_row": 行号
    }
  }
"""
import json
import sys
from pathlib import Path

BASE = Path(__file__).parent
OUTPUT = BASE / '_crawl_data' / 'dai_file_library.json'

from normalize_code import normalize_code


def import_excel(xlsx_path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        print('需要安装 openpyxl: pip install openpyxl')
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Sheet1']

    library = {}
    total = ws.max_row - 1
    with_code = 0
    skipped_no_code = 0
    duplicates = 0  # 戴总 Excel 里自身就有重复的编号

    for i in range(2, ws.max_row + 1):
        filename = ws.cell(i, 1).value or ''
        raw_code = ws.cell(i, 2).value or ''
        raw_code = str(raw_code).strip()

        if not raw_code:
            skipped_no_code += 1
            continue

        with_code += 1
        norm = normalize_code(raw_code)
        if not norm:
            skipped_no_code += 1
            continue

        if norm in library:
            duplicates += 1
            # 多个原始编号都归一到同一个规范化编号，记录所有原始写法
            existing = library[norm].setdefault('all_raw_codes', [library[norm]['raw_code']])
            if raw_code not in existing:
                existing.append(raw_code)
            # 多份 PDF 文件名
            filenames = library[norm].setdefault('all_filenames', [library[norm]['filename']])
            if filename and filename not in filenames:
                filenames.append(filename)
        else:
            library[norm] = {
                'raw_code': raw_code,
                'filename': filename,
                'original_row': i,
            }

    # 保存
    OUTPUT.parent.mkdir(exist_ok=True)
    OUTPUT.write_text(json.dumps(library, ensure_ascii=False, indent=2))

    # 汇报
    print(f'=== 戴总 Excel 导入完成 ===')
    print(f'总行数（除表头）: {total}')
    print(f'  无编号（跳过）: {skipped_no_code}')
    print(f'  有编号: {with_code}')
    print(f'  规范化后唯一编号: {len(library)}')
    print(f'  戴总 Excel 内部重复（多行相同编号）: {duplicates}')
    print(f'输出: {OUTPUT}')

    return library


if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else str(Path.home() / 'Desktop' / '文件库.xlsx')
    import_excel(xlsx)
